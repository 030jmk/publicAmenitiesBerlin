"""publicAmenitiesBerlinTelegramBot.py: Telegram bot which returns the closest public toilets or public water fountains in Berlin, when a location is sent to it"""
__author__      = "Jan Kopankiewicz"


import requests
from openpyxl import load_workbook
import pandas as pd
import kml2geojson
from bs4 import BeautifulSoup
from zipfile import ZipFile
import os
from datetime import datetime
import logging
from telegram import Update, KeyboardButton, ReplyKeyboardMarkup, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, ContextTypes, filters

TOKEN = "YOUR-BOT-TOKEN"

def download_file(url):
    filename = os.path.basename(url)
    try:
        response = requests.get(url, stream=True)
        response.raise_for_status()
        with open(filename, 'wb') as file:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    file.write(chunk)
        print(f"File '{filename}' has been successfully downloaded.")
    except requests.exceptions.RequestException as e:
        print(f"An error occurred while downloading the file: {e}")
    except IOError as e:
        print(f"An error occurred while writing the file: {e}")
def convert_to_float(value):
    if isinstance(value, str):
        return float(value.replace(',', '.'))
    return value

def get_visible_rows(filename, sheet_name):
    wb = load_workbook(filename)
    sheet = wb[sheet_name]
    return [row - 1 for row in range(1, sheet.max_row + 1) if not sheet.row_dimensions[row].hidden]

def find_header_row(df, header_column='Bezirk'):
    header_row = df.index[df.iloc[:, 0] == header_column].tolist()
    return header_row[0] if header_row else None

def haversine_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate the great circle distance between two points on the earth."""
    from math import radians, sin, cos, sqrt, atan2
    R = 6371  # Earth radius in km
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    
    return round(R * c, 2)
def location_cal(df: pd.DataFrame, my_location: tuple[float, float]) -> pd.DataFrame:
    """Calculate distances from a given location to all locations in the DataFrame,
    then return the top 5 closest locations."""
    df['Distance'] = df.apply(lambda row: haversine_distance(my_location[0], my_location[1], 
                                                             row.Breitengrad, row.Laengengrad), axis=1)
    return df.nsmallest(5, 'Distance')

#Berliner Toiletten
def update_toilettes():
    download_file("https://www.berlin.de/sen/uvk/_assets/verkehr/infrastruktur/oeffentliche-toiletten/berliner-toiletten-standorte.xlsx")
    filename = 'berliner-toiletten-standorte.xlsx'
    sheet_name = 'Berlinweit'
    visible_rows = get_visible_rows(filename, sheet_name)

    df_temp = pd.read_excel(filename, sheet_name=sheet_name, header=None, 
                            skiprows=lambda x: x not in visible_rows)

    header_index = find_header_row(df_temp)

    if header_index is not None:
        df_wc = pd.read_excel(filename, sheet_name=sheet_name, 
                            skiprows=lambda x: x not in visible_rows or x < header_index,
                            header=0)  
    else:
        print("Could not find a row starting with 'Bezirk'")
        df_wc = df_temp 

    vertrag_map = {
        1: "Toilettenvertrag mit Wall",
        2: "Pilotprojekt Parktoilettenvertrag",
        3: "Pilottoiletten im Grün/Sonstige öffentliche Toiletten",
        4: "Privat betriebene öffentliche Toiletten" 
    }
    df_wc['Vertrag'] = pd.to_numeric(df_wc['Vertrag'], errors='coerce')
    df_wc['Vertrag'] = df_wc['Vertrag'].map(vertrag_map).fillna(df_wc['Vertrag'])
    df_wc['Breitengrad'] = df_wc['Breitengrad'].apply(convert_to_float)
    df_wc['Laengengrad'] = df_wc['Laengengrad'].apply(convert_to_float)
    return df_wc

#Berliner Wasser

def update_water():
    bwb_url = "https://www.bwb.de/de/trinkbrunnen.php"
    req = requests.get(bwb_url)
    soup = BeautifulSoup(req.content, 'html.parser')
    kmz_url = soup.find("a", class_="trinkbrunnen")['href']
    r = requests.get(kmz_url, allow_redirects=True)
    open('Trinkbrunnen.kmz', 'wb').write(r.content)
    # unzip
    kmz = ZipFile("Trinkbrunnen.kmz", 'r')
    kml = kmz.open('doc.kml', 'r').read()
    # save as xml
    f = open("Trinkbrunnen.xml", "w")
    f.write(str(kml,'utf-8'))
    f.close()
    #convert the xml to more usable json format
    trinkbrunnen_data = kml2geojson.main.convert("Trinkbrunnen.xml", "leaflet")

    rows_list = []
    for i in range(len(trinkbrunnen_data[0]["features"])):
        name=           trinkbrunnen_data[0]["features"][i]["properties"]["name"]
        description =   trinkbrunnen_data[0]["features"][i]["properties"]["description"]
        long, lat =     trinkbrunnen_data[0]["features"][i]["geometry"]["coordinates"][0:2]
        
        dict1 = {'Name': name, 'Description': description, 'Laengengrad': long, 'Breitengrad': lat}
        rows_list.append(dict1)

    # create dataframe 
    return pd.DataFrame.from_dict(rows_list)

def update_police_demo_data():
    """Return a DataFrame containing the demonstrations or protests listed for today including a rough estimation of where it is happening"""
    currentDate = datetime.now().strftime("%d.%m.%Y")
    event_url = "https://www.berlin.de/polizei/service/versammlungsbehoerde/versammlungen-aufzuege/"
    req = requests.get(event_url)
    soup = BeautifulSoup(req.content, 'html.parser')
    table_r = soup.findAll("tr", class_="odd line_1")
    demo_list = []
    for row in table_r:
        try:
            if row.find("td", class_="text", headers="Datum").text == currentDate:
                Datum = row.find("td", class_="text", headers="Datum").text
                Von = row.find("td", class_="text", headers="Von").text
                Bis = row.find("td", class_="text", headers="Bis").text
                Thema = row.find("td", class_="text", headers="Thema").text
                PLZ = int(row.find("td", class_="text", headers="PLZ").text)
                Versammlungsort = row.find("td", class_="text", headers="Versammlungsort").text
                Aufzugsstrecke = row.find("td", class_="text", headers="Aufzugsstrecke").text
                demo_list.append([Datum,Von,Bis,Thema,PLZ,Versammlungsort,Aufzugsstrecke])
        except:
            continue
    demo_df = pd.DataFrame(demo_list, columns=["Datum","Von","Bis","Thema","PLZ","Versammlungsort","Aufzugsstrecke"])
    demo_df['Breitengrad'] = demo_df['PLZ'].map(lambda plz: plz_map.get(plz, (None, None))[1])
    demo_df['Laengengrad'] = demo_df['PLZ'].map(lambda plz: plz_map.get(plz, (None, None))[0])
    return demo_df.sort_values('Von')


# Postleitzahl Map
plz_map = {10115: (13.384607458757577, 52.53225310749616),
 10119: (13.405320915304019, 52.53047717452423),
 10117: (13.38722234829472, 52.516965560388),
 10178: (13.409628466424998, 52.52131243046372),
 10179: (13.416335442222064, 52.51219339888968),
 10243: (13.439382729365533, 52.51230622328652),
 10245: (13.464755384582356, 52.50065849625331),
 10247: (13.46555361594058, 52.516159537561606),
 10249: (13.442772448085373, 52.52376255730941),
 10315: (13.514764286811292, 52.51322993265686),
 10317: (13.490766865201651, 52.49790145501775),
 10318: (13.52868724331732, 52.48348496684792),
 10319: (13.51880237290777, 52.499193681905616),
 10365: (13.496861997511203, 52.52061304615122),
 10367: (13.482099234822453, 52.5246228538931),
 10369: (13.46944982046535, 52.529475650969495),
 10405: (13.42570376752267, 52.535182312912546),
 10407: (13.449170806424025, 52.53360703864684),
 10409: (13.441357051572137, 52.54431847498743),
 10435: (13.411186295587143, 52.53776375845449),
 10437: (13.412580850688249, 52.544853189822),
 10439: (13.412114584540683, 52.55215950650852),
 10551: (13.337163496142006, 52.530720056080916),
 10553: (13.321465483262193, 52.5305064491924),
 10555: (13.33545705220396, 52.52153170911194),
 10557: (13.35943703868717, 52.52332350552952),
 10559: (13.349920246972744, 52.53012311902867),
 10585: (13.305687562670663, 52.51519648047775),
 10587: (13.319516472927122, 52.51844731471893),
 10589: (13.305709015285817, 52.527550277121854),
 10623: (13.327363836111795, 52.50882403120826),
 10625: (13.314686440987971, 52.50945814402798),
 10627: (13.302995552558253, 52.50798435303231),
 10629: (13.308587338266317, 52.502795190278626),
 10707: (13.313752951601778, 52.49665671691031),
 10709: (13.303116608927256, 52.49388187834184),
 10711: (13.290451365138114, 52.49812112980343),
 10713: (13.3132725311462, 52.4850887035849),
 10715: (13.328877316645858, 52.48244496141088),
 10717: (13.327547835159514, 52.490797857138496),
 10719: (13.325678619451482, 52.49884625285831),
 10777: (13.342702106333128, 52.49745514955069),
 10779: (13.339475392967456, 52.492111831434606),
 10781: (13.352914193132012, 52.4935684258165),
 10783: (13.36237019394093, 52.4964239900215),
 10785: (13.36424990258966, 52.507309547686845),
 10787: (13.34386932460584, 52.50777991902037),
 10789: (13.337703334254424, 52.50166742652484),
 10823: (13.350881278556777, 52.487308839107946),
 10825: (13.341243634796259, 52.483759578982635),
 10827: (13.354257774407076, 52.48377647229864),
 10829: (13.360796518251403, 52.47618810975703),
 10961: (13.39747069383011, 52.49262275989204),
 10963: (13.381258333215879, 52.50016090220396),
 10965: (13.39488499701448, 52.485375215726236),
 10967: (13.416415039260794, 52.49050264888512),
 10969: (13.401131913828474, 52.5024880493976),
 10997: (13.43555805758727, 52.50092154478418),
 10999: (13.426558517079119, 52.49691730607814),
 12043: (13.437059652549424, 52.47989834134753),
 12045: (13.4392318279728, 52.48546376140066),
 12047: (13.428474510622515, 52.49052454933148),
 12049: (13.422009717649571, 52.476348668074564),
 12051: (13.42987658460868, 52.46690101284063),
 12053: (13.432528934707053, 52.476838288722455),
 12055: (13.448598746173031, 52.471208546855706),
 12057: (13.463283331971413, 52.46839875771354),
 12059: (13.451286098618839, 52.48091987610456),
 12099: (13.4023349994847, 52.46440180369397),
 12101: (13.379067939668047, 52.47849483618643),
 12103: (13.374691679393424, 52.46405527434142),
 12105: (13.371378239048981, 52.4492182341782),
 12107: (13.39169561926881, 52.4312233429101),
 12109: (13.399354742820034, 52.446437669268846),
 12157: (13.346192882606559, 52.46531937874084),
 12159: (13.33691777492158, 52.47367758592234),
 12161: (13.326968406690545, 52.470378554493706),
 12163: (13.31845865212027, 52.462641233568114),
 12165: (13.314838153232584, 52.455665181867445),
 12167: (13.333792145887184, 52.44859425042259),
 12169: (13.343532831550357, 52.45479056786636),
 12203: (13.309548541854562, 52.4443763196241),
 12205: (13.29452340895023, 52.43397281284941),
 12207: (13.313201465002296, 52.41988650094883),
 12209: (13.329097359700793, 52.41791658317226),
 12247: (13.346216565231918, 52.43947485499027),
 12249: (13.351813165344238, 52.42636551712031),
 12277: (13.375032628919524, 52.41339546957715),
 12279: (13.353053044314894, 52.410626611040534),
 12305: (13.402072797096357, 52.40326994732729),
 12307: (13.390696838216329, 52.388629911251684),
 12309: (13.417145044045018, 52.39048701051443),
 12347: (13.428134102720877, 52.450870338262156),
 12349: (13.422080228855206, 52.42525449703233),
 12351: (13.455512684667102, 52.43275828308654),
 12353: (13.458921988798815, 52.42273774313897),
 12355: (13.497828066401619, 52.41099140301207),
 12357: (13.490523118452181, 52.42930017882094),
 12359: (13.453134536462507, 52.44733396802487),
 12435: (13.467183898610612, 52.48655920688828),
 12437: (13.48168105737062, 52.46239586702063),
 12459: (13.528082371587278, 52.46556876524206),
 12487: (13.505149578645996, 52.443705865252674),
 12489: (13.543153325823347, 52.43560426111443),
 12524: (13.541654222271243, 52.412832884792735),
 12526: (13.564209625978904, 52.397638858747065),
 12527: (13.633882041175635, 52.38562496072207),
 12555: (13.579098294658408, 52.46267360803515),
 12557: (13.59175495071152, 52.430343521801944),
 12559: (13.663272947067718, 52.41489702770125),
 12587: (13.636164296385875, 52.458611027802064),
 12589: (13.703360765641127, 52.4438132942781),
 12619: (13.58829149686148, 52.52348896949066),
 12621: (13.587807754058906, 52.50272614501193),
 12623: (13.616493890819186, 52.502591371726446),
 12627: (13.613493972069534, 52.53722540813074),
 12629: (13.590114893983868, 52.54131146630923),
 12679: (13.565985347868985, 52.550137779217394),
 12681: (13.536691577327467, 52.53690378944878),
 12683: (13.559059198265544, 52.50751934032599),
 12685: (13.565008324791537, 52.53908677744801),
 12687: (13.564473457443823, 52.55641340047163),
 12689: (13.56751600687462, 52.566476173013776),
 13051: (13.490844938375092, 52.581510242877),
 13053: (13.504599687469637, 52.550014431889664),
 13055: (13.495996562275451, 52.540085051396055),
 13057: (13.541474255141146, 52.57105312774011),
 13059: (13.521691039944386, 52.58085224331766),
 13086: (13.448184825048454, 52.556479096011834),
 13088: (13.470798827594919, 52.560323494890454),
 13089: (13.44099736024904, 52.57068024201811),
 13125: (13.482939796838133, 52.63285914727104),
 13127: (13.43803639424625, 52.619999928293986),
 13129: (13.457927558520657, 52.59205784915984),
 13156: (13.399679345417574, 52.58235941503544),
 13158: (13.383482175401902, 52.59319891839218),
 13159: (13.39781930738672, 52.622980185622865),
 13187: (13.408406728535377, 52.569544289010025),
 13189: (13.421922791184404, 52.56428133261613),
 13347: (13.365455385862616, 52.54906035716649),
 13349: (13.347338321717679, 52.55798834244686),
 13351: (13.33282903404662, 52.55065264364231),
 13353: (13.349493150452554, 52.54159282020101),
 13355: (13.390584510139545, 52.54177403402419),
 13357: (13.382545290564279, 52.55025478348207),
 13359: (13.38508861070656, 52.55987577586028),
 13403: (13.322394750985106, 52.57390962656044),
 13405: (13.296715655546139, 52.55956271601847),
 13407: (13.351152712066776, 52.5726565884841),
 13409: (13.37136141499762, 52.567874990707),
 13435: (13.345583555610988, 52.602047574882135),
 13437: (13.328433400733003, 52.5904606987358),
 13439: (13.35836552281583, 52.597633923946645),
 13465: (13.289551932921368, 52.639894021363325),
 13467: (13.307476993839325, 52.61710500376173),
 13469: (13.34217007480002, 52.611886124665894),
 13503: (13.248754734738577, 52.61216227960929),
 13505: (13.240436851292232, 52.5839010321919),
 13507: (13.271706478682745, 52.576502619970846),
 13509: (13.300585116675384, 52.589187536239734),
 13581: (13.179372234866618, 52.53102430675617),
 13583: (13.182353354198945, 52.54365970632263),
 13585: (13.204912806155, 52.54772720782926),
 13587: (13.185425730622057, 52.57671807096514),
 13589: (13.167559316007088, 52.55702798271659),
 13591: (13.140451047305051, 52.53446995910025),
 13593: (13.167205216173901, 52.51482610221952),
 13595: (13.19622459336038, 52.51161478298646),
 13597: (13.219494683751504, 52.52724736599811),
 13599: (13.23500664599951, 52.54629375977735),
 13627: (13.299091755015624, 52.53982806674212),
 13629: (13.266121838408706, 52.54217370067193),
 14050: (13.268338227370933, 52.52082672608213),
 14052: (13.256858799554047, 52.51558920656036),
 14053: (13.238696178740382, 52.515905226931764),
 14193: (13.236512191039548, 52.48312900126494),
 14057: (13.287913704517218, 52.507250172456125),
 14059: (13.287773930793634, 52.52052390612372),
 14089: (13.151643265502072, 52.47078498219171),
 14109: (13.143986703427162, 52.41973521998986),
 14129: (13.202578358463798, 52.446286478934425),
 14163: (13.238505101240266, 52.43683095261234),
 14165: (13.253558879540764, 52.41751915632142),
 14167: (13.276466893449365, 52.42117084955016),
 14169: (13.257318230890549, 52.44961801437708),
 14195: (13.282867034390405, 52.458882967786316),
 14197: (13.311789699628607, 52.47335862767753),
 14199: (13.29507096395517, 52.477661074858325),
 12439: (13.528644321780456, 52.45277028153434),
 14055: (13.244731875810201, 52.501950954930614),
 15537: (13.687388529541318, 52.38570793162978),
 15566: (13.705366666666666, 52.459782999999995),
 15569: (13.756057865644234, 52.445953729962476)}

# Enable logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO
)
logger = logging.getLogger(__name__)
# Global variables for dataframes
toilette_df = None
water_df = None
demo_df = None

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Send message on `/start`."""
    user = update.message.from_user
    logger.info("User %s started the conversation.", user.first_name)
    send_location_keyboard = [[KeyboardButton(text="Send current location", request_location=True)]]
    await update.message.reply_text('Help me figure out where you are by sending me your location.',
        reply_markup=ReplyKeyboardMarkup(send_location_keyboard, one_time_keyboard=False))

async def button(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.message.from_user
    user_location = update.message.location

    choice_keyboard = [
        [InlineKeyboardButton("Public Toilet", callback_data=f"wc,{user_location.latitude},{user_location.longitude}")],
        [InlineKeyboardButton("Potable Water", callback_data=f"water,{user_location.latitude},{user_location.longitude}")],
        [InlineKeyboardButton("Demonstrations", callback_data=f"demo,{user_location.latitude},{user_location.longitude}")],
        [InlineKeyboardButton("Update Lists", callback_data="update")]
    ]

    await update.message.reply_text("Which public amenity are you looking to find?",
                                    reply_markup=InlineKeyboardMarkup(choice_keyboard))

async def pick_one(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    global toilette_df, water_df, demo_df
    query = update.callback_query
    await query.answer()

    if query.data == "update":
        demo_df = update_police_demo_data()
        toilette_df = update_toilettes() 
        water_df = update_water()
        await query.message.reply_text("Lists are up to date")
        return

    choice, lat, lon = query.data.split(',')
    my_location = (float(lat), float(lon))

    if choice == "wc":
        Top5 = location_cal(toilette_df, my_location)
        # Check if 'Description' column exists, if not use an alternative
        description_column = 'Description' if 'Description' in Top5.columns else 'Standort'
        keyboard = [
            [InlineKeyboardButton(text=f"{Top5['Distance'].iloc[i]:.2f}km - {Top5[description_column].iloc[i]}",
                                  url=f"https://maps.apple.com/maps?q={Top5['Breitengrad'].iloc[i]},{Top5['Laengengrad'].iloc[i]}")]
            for i in range(min(3, len(Top5)))
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.reply_text("Closest Options:", reply_markup=reply_markup)

    elif choice == "water":
        Top5w = location_cal(water_df, my_location)
        keyboard = [
            [InlineKeyboardButton(text=f"{Top5w['Distance'].iloc[i]}km - {Top5w['Name'].iloc[i]}", 
                                  url=f"https://maps.apple.com/maps?q={Top5w['Breitengrad'].iloc[i]},{Top5w['Laengengrad'].iloc[i]}")]
            for i in range(3)
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.reply_text("Closest Options:", reply_markup=reply_markup)

    elif choice == "demo":
        Top5demo = location_cal(demo_df, my_location)
        keyboard = [
            [InlineKeyboardButton(text=f"~{round(Top5demo['Distance'].iloc[i],1)}km - {Top5demo['Thema'].iloc[i]}", 
                                  url=f"https://maps.apple.com/maps?q={Top5demo['Versammlungsort'].iloc[i]},{Top5demo['PLZ'].iloc[i]} Berlin")]
            for i in range(3)
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.reply_text("Nearby Protests:", reply_markup=reply_markup)


async def update_data():
    global toilette_df, water_df, demo_df
    demo_df = update_police_demo_data()
    toilette_df = update_toilettes() 
    water_df = update_water()



def main() -> None:
    """Run the bot."""
    # Create the Application and pass it your bot's token.
    application = Application.builder().token(TOKEN).build()

    # Add handlers
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CallbackQueryHandler(pick_one))
    application.add_handler(MessageHandler(filters.LOCATION, button))

    # Run the bot until the user presses Ctrl-C
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()


